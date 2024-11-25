import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import plotly.graph_objects as go
import time
import os
import pickle

from plotly.subplots import make_subplots
from my_lib.utility import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_report(n_poblacion, poblacion, time=None):
    info = np.zeros((8))
    info[0] = n_poblacion
    info[1] = np.around(np.mean(poblacion[:, 3]), decimals=2)
    info[2] = np.around(np.var(poblacion[:, 3]), decimals=2)
    info[3] = np.around(np.std(poblacion[:, 3]), decimals=2)
    info[4] = np.around(np.min(poblacion[:, 3]), decimals=2)
    info[5] = np.around(np.max(poblacion[:, 3]), decimals=2)
    info[6] = np.around(diversity_rate(poblacion), decimals=2)
    if time:
        info[7] = time

    return info


def intial_stats(pobs, msg):
    reportes_iniciales = np.array([get_report(i + 1, p) for i, p in enumerate(pobs)])
    # Comportamiento de las poblaciones iniciales
    fig = make_subplots(
        rows=2,
        cols=3,
        subplot_titles=[
            "Costo promedio",
            "Varianza del costo",
            "Desviación estándar",
            "Individuos con menor costo población inicial",
            "Individuos mayor costo población inicial",
            "Diversidad de genes",
        ],
    )

    fig.add_trace(
        go.Scatter(
            x=reportes_iniciales[:, 0],
            y=reportes_iniciales[:, 1],
            mode="markers+lines",
            name="Costo promedio inicial",
            marker=dict(symbol="star-diamond", size=8),
        ),
        row=1,
        col=1,
    )
    fig.add_trace(
        go.Scatter(
            x=reportes_iniciales[:, 0],
            y=reportes_iniciales[:, 2],
            mode="markers+lines",
            marker=dict(symbol="diamond-x", size=8),
            name="Varianza costo",
        ),
        row=1,
        col=2,
    )
    fig.add_trace(
        go.Scatter(
            x=reportes_iniciales[:, 0],
            y=reportes_iniciales[:, 3],
            mode="markers+lines",
            marker=dict(symbol="octagon", size=8),
            name="Desviación costo",
        ),
        row=1,
        col=3,
    )
    fig.add_trace(
        go.Scatter(
            x=reportes_iniciales[:, 0],
            y=reportes_iniciales[:, 4],
            mode="markers+lines",
            marker=dict(symbol="bowtie", size=8),
            name="Ruta con Menor Distancia",
        ),
        row=2,
        col=1,
    )
    fig.add_trace(
        go.Scatter(
            x=reportes_iniciales[:, 0],
            y=reportes_iniciales[:, 5],
            mode="markers+lines",
            marker=dict(symbol="bowtie", size=8),
            name="Ruta con Mayor Distancia",
        ),
        row=2,
        col=2,
    )
    fig.add_trace(
        go.Scatter(
            x=reportes_iniciales[:, 0],
            y=reportes_iniciales[:, 6],
            mode="markers+lines",
            marker=dict(symbol="hexagram", size=8),
            name="Costo mayor",
        ),
        row=2,
        col=3,
    )

    fig.update_layout(height=800, width=1600)  # Ajusta el tamaño de la figura
    # Agrega títulos a los ejes x y y de cada gráfica
    fig.update_layout(
        title=f"{msg}",
        xaxis1=dict(title="número población"),  # Eje x de la primera gráfica
        yaxis1=dict(title="Costo"),  # Eje y de la primera gráfica
        xaxis2=dict(title="número población"),  # Eje x de la segunda gráfica
        yaxis2=dict(title="Costo"),  # Eje y de la segunda gráfica
        xaxis3=dict(title="número población"),  # Eje x de la tercera gráfica
        yaxis3=dict(title="Costo"),  # Eje y de la tercera gráfica
        xaxis4=dict(title="número población"),  # Eje x de la tercera gráfica
        yaxis4=dict(title="Costo"),  # Eje y de la tercera gráfica
        xaxis5=dict(title="número población"),  # Eje x de la tercera gráfica
        yaxis6=dict(title="Genes"),  # Eje y de la tercera gráfica
    )
    fig.show()


def get_poblaciones_iniciales(
    pob_dir,
    ciudades,
    funcion_evaluar,
    funcion_validar=None,
    n=5,
    tam_pob=100,
):
    """
    Genera y carga las poblaciones iniciales para un algoritmo genético.

    Si el archivo de poblaciones no existe, la función lo crea generando las poblaciones iniciales y guardándolas en el archivo.
    Si el archivo ya existe, la función lo carga y devuelve las poblaciones.

    Parámetros:
        pob_dir(str): Ruta en donde se almacenará el archivo de poblaciones.
        ciudades(matrix): Matriz de ciudades con su id y nombre.
        funcion_evaluar(función): Función para evaluar la aptitud de un individuo
        funcion_validar(función): Función para validar un cromosoma por defecto no hay
        n (int): Número de poblaciones iniciales a generar (por defecto, 5).
        tam_pob (int): Tamaño de cada población (por defecto, 60).

    Valor de retorno:
        list: Lista de poblaciones iniciales, donde cada población es una matriz de numpy con 4 columnas: ruta, aptitud, individuo y aptitud.
    """
    poblaciones = []
    file_name = f"{n}pobs_{tam_pob}tam.pkl"
    POB_PATH = os.path.join(pob_dir, file_name)
    if not os.path.exists(POB_PATH):
        print("El archivo de poblaciones no existse, creando archivo")

        for i in range(n):
            recorrido = generar_poblacion_perm(ciudades, tam_pob)

            poblacion = np.zeros((recorrido.shape[0], 4), dtype="object")
            if funcion_validar is None:
                poblacion[:, 0] = [c.astype(int) for c in recorrido]  # Ruta
            else:
                poblacion[:, 0] = [
                    funcion_validar(c.astype(int)) for c in recorrido
                ]  # Ruta
            poblacion[:, 1] = np.ones((recorrido.shape[0]))
            poblacion[:, 2] = [
                " ".join([chr(c + 64) for c in perm.astype(int)]) for perm in recorrido
            ]
            # poblacion[:, 2] = [individuo_toString(c.astype(int)) for c in recorrido]
            # poblacion[:, 3] = np.zeros((recorrido.shape[0]))  # Aptitud
            poblacion[:, 3] = funcion_evaluar(poblacion)

            poblaciones.append(poblacion)

        try:
            with open(POB_PATH, "wb") as f:
                pickle.dump(poblaciones, f)
            print("Archivo de poblaciones creado correctamente")
        except Exception as e:
            print(f"Error al guardar archivo: {e}")

    else:
        try:
            with open(POB_PATH, "rb") as f:
                poblaciones = pickle.load(f)
            print("Archivo de poblaciones cargado correctamente")
        except Exception as e:
            print(f"Error al cargar archivo: {e}")

    return poblaciones


def historial_cruces(descendencia):
    hijos = [ind[0] for ind in descendencia]
    padres_codificados = [ind[1] for ind in descendencia]
    padres_codificados = [
        [p.replace(" ", "") for p in padres] for padres in padres_codificados
    ]
    padres_decodificados = [
        [[ord(c) - 64 for c in p] for p in padres] for padres in padres_codificados
    ]

    historial = []

    for j in range(0, len(hijos), 2):
        fila = np.array(
            [
                padres_decodificados[j][0],
                padres_decodificados[j][1],
                hijos[j],
                hijos[j + 1],
            ]
        )
        historial.append(fila)

    return np.array(historial)


def show_report(resutaldos, n_gen, total_pobs=5, msg=""):
    gen_reports = []
    for i, rg in enumerate(resutaldos):
        rg_tiempos = rg[:, 1]
        rg_res = rg[:, 0]

        rg_res_pobs = np.array([respob[0] for respob in rg_res])
        rg_report = np.array(
            [
                get_report(k + 1, rpob, rg_tiempos[k])
                for k, rpob in enumerate(rg_res_pobs)
            ]
        )
        gen_reports.append(rg_report)

    fig = make_subplots(
        rows=3,
        cols=2,
        subplot_titles=[
            "Costo promedio",
            "Varianza",
            "Desviación estándar",
            "Ruta con la Menor Distancia",
            "Ruta con la Mayor Distancia",
            "Diversidad genética",
        ],
    )

    colors = [
        "#f5bde6",  # Pink
        "#c6a0f6",  # Mauve
        "#ed8796",  # Red
        "#a6da95",  # Green
        "#7dc4e4",  # Sapphire
        "#8bd5ca",  # Teal
        "#f5a97f",  # Peach
        "#91d7e3",  # Sky
        "#8aadf4",  # Blue
        "#5b6078",  # Surface
    ]

    marker_symbols = [
        "pentagon",
        "diamond",
        "circle",
        "hexagram",
        "star",
        "square",
        "x",
        "star-square",
        "bowtie",
        "circle-x",
    ]
    for i, pop in enumerate(gen_reports):
        fig.add_trace(
            go.Scatter(
                x=pop[:, 0],
                y=pop[:, 1],
                name=f"Población {i+1}",
                mode="markers+lines",
                marker=dict(symbol=marker_symbols[i], size=8, color=colors[i]),
                legendgroup=f"Población {i+1}",
            ),
            row=1,
            col=1,
        )

        fig.add_trace(
            go.Scatter(
                x=pop[:, 0],
                y=pop[:, 2],
                name=f"Población {i+1}",
                mode="markers+lines",
                marker=dict(symbol=marker_symbols[i], size=8, color=colors[i]),
                legendgroup=f"Población {i+1}",
            ),
            row=1,
            col=2,
        )

        fig.add_trace(
            go.Scatter(
                x=pop[:, 0],
                y=pop[:, 3],
                name=f"Población {i+1}",
                mode="markers+lines",
                marker=dict(symbol=marker_symbols[i], size=8, color=colors[i]),
                legendgroup=f"Población {i+1}",
            ),
            row=2,
            col=1,
        )

        fig.add_trace(
            go.Scatter(
                x=pop[:, 0],
                y=pop[:, 4],
                name=f"Población {i+1}",
                mode="markers+lines",
                marker=dict(symbol=marker_symbols[i], size=8, color=colors[i]),
                legendgroup=f"Población {i+1}",
            ),
            row=2,
            col=2,
        )

        fig.add_trace(
            go.Scatter(
                x=pop[:, 0],
                y=pop[:, 5],
                name=f"Población {i+1}",
                mode="markers+lines",
                marker=dict(symbol=marker_symbols[i], size=8, color=colors[i]),
                legendgroup=f"Población {i+1}",
            ),
            row=3,
            col=1,
        )

        fig.add_trace(
            go.Scatter(
                x=pop[:, 0],
                y=pop[:, 6],
                name=f"Población {i+1}",
                mode="markers+lines",
                marker=dict(symbol=marker_symbols[i], size=8, color=colors[i]),
                legendgroup=f"Población {i+1}",
            ),
            row=3,
            col=2,
        )

    for i in range(6):
        fig.update_xaxes(title_text="Generación", row=(i // 2) + 1, col=(i % 2) + 1)

    fig.update_layout(height=1200, width=1500, title=f"{msg}")
    fig.show()


def save_results(res_file_name, reslt, pob_dir):
    res_path = os.path.join(pob_dir, f"{res_file_name}.pkl")
    try:
        with open(res_path, "wb") as f:
            pickle.dump(reslt, f)
    except Exception as e:
        print(f"Error al gaurdar el archvio de resultados: {e}")


def create_excel_hist(reslt, file_name, pob_dir):
    # Colores de resaltado
    fill_green = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    fill_yellow = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    # arch = "exp1_no_mutacion_style2.xlsx"
    hist_file = os.path.join(pob_dir, f"{file_name}.xlsx")
    # hist_file = f"../files/act_2/historial/{file_name}.xlsx"

    with pd.ExcelWriter(hist_file, engine="openpyxl") as writer:
        for i, res in enumerate(reslt):
            start_row = 0

            for g, gen_res in enumerate(res[:, 0]):
                # Crear los DataFrames
                historial_generaciones = np.zeros(
                    (gen_res[0][:, 0].shape[0], 2), dtype="object"
                )
                historial_generaciones[:, 0] = np.array(
                    [" ".join([str(c) for c in r]) for r in gen_res[0][:, 0]]
                )
                historial_generaciones[:, 1] = gen_res[0][:, 3]

                hg_df = pd.DataFrame(
                    historial_generaciones, columns=["Ruta", "Distancia"]
                )

                # DataFrame de cruces con columnas adicionales
                historial_cruces = np.zeros((gen_res[1].shape[0], 4), dtype="object")
                historial_cruces[:, 0] = np.array(
                    [" ".join([str(c) for c in r]) for r in gen_res[1][:, 0]]
                )
                historial_cruces[:, 1] = np.array(
                    [" ".join([str(c) for c in r]) for r in gen_res[1][:, 1]]
                )
                historial_cruces[:, 2] = np.array(
                    [" ".join([str(c) for c in r]) for r in gen_res[1][:, 2]]
                )
                historial_cruces[:, 3] = np.array(
                    [" ".join([str(c) for c in r]) for r in gen_res[1][:, 3]]
                )

                hc_df = pd.DataFrame(
                    historial_cruces, columns=["Padre 1", "Padre 2", "Hijo 1", "Hijo 2"]
                )
                hc_df["Apto h1"] = hc_df["Hijo 1"].isin(hg_df["Ruta"])
                hc_df["Apto h2"] = hc_df["Hijo 2"].isin(hg_df["Ruta"])

                # DataFrame de mutantes con columna adicional
                historial_mutante = gen_res[2]
                hm_df = pd.DataFrame(
                    historial_mutante,
                    columns=["Original", "Mutante", "Incio mutación", "Genes mutados"],
                )
                hc_df["mutado 1"] = hc_df["Hijo 1"].isin(hm_df["Original"])
                hc_df["mutado 2"] = hc_df["Hijo 2"].isin(hm_df["Original"])
                hm_df["Apto mut"] = hm_df["Mutante"].isin(hg_df["Ruta"])

                # Agregar columnas "Cruza" y "Mutado" a hg_df
                hg_df["Cruza"] = hg_df["Ruta"].isin(hc_df["Hijo 1"]) | hg_df[
                    "Ruta"
                ].isin(hc_df["Hijo 2"])
                hg_df["Mutado"] = hg_df["Ruta"].isin(hm_df["Mutante"])

                start_col = 0
                hg_df.to_excel(
                    writer,
                    sheet_name=f"Poblacion_{i}",
                    index=False,
                    startrow=start_row,
                    startcol=start_col,
                )
                start_col += hg_df.shape[1] + 1

                hc_df.to_excel(
                    writer,
                    sheet_name=f"Poblacion_{i}",
                    index=False,
                    startrow=start_row,
                    startcol=start_col,
                )
                start_col += hc_df.shape[1] + 1

                hm_df.to_excel(
                    writer,
                    sheet_name=f"Poblacion_{i}",
                    index=False,
                    startrow=start_row,
                    startcol=start_col,
                )

                start_row += hg_df.shape[0] + 2

    # Resaltar las celdas
    wb = load_workbook(hist_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            ruta_value = row[0].value  # Valor de la columna "Ruta"
            cruza_value = row[2].value  # Valor de la columna "Cruza"
            mutado_value = row[3].value  # Valor de la columna "Mutado"

            # Resaltar en verde si "Cruza" es True
            if cruza_value:
                row[0].fill = fill_green
            # Resaltar en amarillo si "Mutado" es True
            if mutado_value:
                row[0].fill = fill_yellow

    wb.save(hist_file)


def seleccion_cruza(poblacion, op_seleccion=0, op_cruza=3, mode="min"):
    desc = get_next_gen(poblacion, op_seleccion, op_cruza, mode)
    cruces = historial_cruces(desc)
    hijos = np.array([hijo[0] for hijo in desc])
    # new_age = poblacion.copy()
    new_age = np.zeros((hijos.shape[0], 4), dtype="object")
    for i, hijo in enumerate(hijos):
        # nuevo_indiviudo = np.zeros((1, 4), dtype="object")
        # nuevo_indiviudo[0, 0] = hijo
        # nuevo_indiviudo[0, 1] = 1.0
        # nuevo_indiviudo[0, 2] = " ".join([chr(c + 64) for c in hijo])
        # nuevo_indiviudo[0, 3] = evaluar(nuevo_indiviudo)[0]
        # new_age = np.vstack([new_age, nuevo_indiviudo])
        new_age[i, 0] = hijo
        new_age[i, 1] = 1.0
        new_age[i, 2] = " ".join([chr(c + 64) for c in hijo])
    # new_age[:, 3] = evaluar(new_age)
    return new_age, cruces


def msg_info_exp(msg):
    return (
        "===================================================================\n"
        + f"{msg}\n"
    )


def msg_info_inicial(num_pob, poblacion):
    info = "===================================================================\n"
    info += f"Población inicial #{num_pob + 1}\n"
    info += f"Costo promedio inicial {np.mean(poblacion[:, 3])}\n"
    return info


def msg_info_generacion(num_gen, poblacion, tiempo):
    info = "...................................................................\n"
    info += f"Generación: {num_gen}\n"
    info += f"Costo promedio: {np.mean(poblacion[:, 3])}\n"
    info += f"Varianza del costo en la generación: {np.var(poblacion[:, 3])}\n"
    info += (
        f"Desviación estándar del costo en la generación: {np.std(poblacion[:, 3])}\n"
    )
    info += f"Diversidad promedio: {diversity_rate(poblacion)} genes\n"
    info += f"Menor costo de toda la generación: {np.min(poblacion[:, 3])}\n"
    info += f"Mayor costo de toda la generación: {np.max(poblacion[:, 3])}\n"
    info += f"Tiempo de ejecución de la generación {num_gen}: {tiempo}\n"
    info += "...................................................................\n"
    return info


def msg_tiempo_pob(num_pob, tiempo):
    info = "--------------------------------------------------------------------\n"
    info += f"Tiempo de ejecución para la población #{num_pob + 1}: {tiempo}\n"
    info += "===================================================================\n"
    return info
